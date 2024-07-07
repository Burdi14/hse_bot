import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
url_main = 'https://ba.hse.ru/base2024'
import openpyxl

def main():
    snils = '180-937--'
    r = requests.get(url_main)
    with open('main_page.html', 'wb') as f:
        f.write(r.content)

    with open('main_page.html', 'r') as file_page:
        page = file_page.read()
        soup = bs(page, 'lxml')
        table = soup.find('tbody')
        programs = table.find_all('tr')[1:]

        program_url = []
        for program in programs:
            link = program.find('a')['href']

            r = requests.get(link)
            with open('cur.xlsx', 'bw') as f:
                f.write(r.content)

            wb = openpyxl.load_workbook('cur.xlsx')
            sheet = wb.active

            row = 0
            program_name = sheet['A1'].value
            free_places = 0
            commercial_places = 0
            snils_state = ''
            snils_rate = 0
            rating = []
            found = 0
            scan_rate = 0
            for row in range(1, sheet.max_row + 1):
                row += 1
                if scan_rate:
                    if sheet.cell(row, 2).value == snils:
                        found = 1
                        print(program_name)
                        print('Your points: ', sheet.cell(row, 16).value)
                        snils_rate = int(sheet.cell(row, 16).value)
                        snils_state = sheet.cell(row, 17).value
                    if sheet.cell(row, 16).value != None:
                        rating.append(int(sheet.cell(row, 16).value))
                    else:
                        rating.append(0)
                else:
                    if row == 2:
                        program_name = sheet.cell(row, 1).value
                        continue
                    if row == 8:
                        free_places = []
                        for col in range(2, sheet.max_column + 1):
                            free_places.append(sheet.cell(row, col).value)
                        for place in free_places:
                            if place != None:
                                free_places = int(place)
                                break
                        continue
                    if row == 10:
                        commercial_places= []
                        for col in range(2, sheet.max_column + 1):
                            commercial_places.append(sheet.cell(row, col).value)
                        for place in commercial_places:
                            if place != None:
                                commercial_places= int(place)
                                break
                        continue

                    if sheet.cell(row, 1).value != None and str(sheet.cell(row, 1).value)[0] in '1234567890':
                        scan_rate = 1
                        if sheet.cell(row, 2).value == snils:
                            found = 1
                            print('your points: ', sheet.cell(row, 16).value, end='')
                            snils_rate = int(sheet.cell(row, 16).value)
                            snils_state = sheet.cell(row, 17).value
                        if sheet.cell(row, 16).value != None:
                            rating.append(int(sheet.cell(row, 16).value))
                        else:
                            rating.append(0)
            if found:
                rating.sort(reverse=True)
                position_1 = rating.index(snils_rate)
                rating.reverse()
                position_2 = len(rating) - rating.index(snils_rate)
            #return {position_1, position_2, free_places, commercial_places}
                print('Positions:  ', position_1,'-', position_2, sep = '')
                print('Free places: ',free_places, '\nCommercial places: ', commercial_places, '\n')

            wb.close()

if __name__ == '__main__':
    main()


