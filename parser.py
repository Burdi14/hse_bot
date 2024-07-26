import requests
from bs4 import BeautifulSoup as bs
import openpyxl

url_main = 'https://ba.hse.ru/base2024'

def get_main_page(url):
    r = requests.get(url)
    with open('main_page.html', 'wb') as f:
        f.write(r.content)

def get_programs():
    get_main_page(url_main)
    with open('main_page.html', 'r') as file_page:
        page = file_page.read()
        soup = bs(page, 'lxml')
        table = soup.find('tbody')
        program_links = [program.find('a')['href'] for program in table.find_all('tr')[1:]]
        return program_links
def parse_program(link, snils):
    snils = snils[:3]+'-'+snils[3:6]+'-'+snils[6:9]+' ' +snils[9:]
    r = requests.get(link)
    with open('cur.xlsx', 'bw') as f:
        f.write(r.content)

    wb = openpyxl.load_workbook('cur.xlsx')
    sheet = wb.active

    row = 0
    program_name = sheet['A1'].value
    free_places = 0
    commercial_places = 0
    snils_state = ''
    snils_rate = 0
    rating = []
    rating_with_docs = []
    found = 0
    scan_rate = 0
    for row in range(1, sheet.max_row + 1):
        row += 1
        if scan_rate:
            if sheet.cell(row, 2).value == snils:
                found = 1
                snils_rate = int(sheet.cell(row, 18).value)
                snils_state = sheet.cell(row, 19).value
            if sheet.cell(row, 18).value != None:
                rate = 0
                if sheet.cell(row, 3).value == 'Да':
                    rate = 310
                else:
                    rate = int(sheet.cell(row, 18).value)
                rating.append(rate)
                if sheet.cell(row, 20).value == 'Да':
                    rating_with_docs.append(rate)
            else:
                rating.append(0)
        else:
            if row == 2:
                program_name = sheet.cell(row, 1).value
                continue
            if row == 8:
                free_places = []
                for col in range(2, sheet.max_column + 1):
                    free_places.append(sheet.cell(row, col).value)
                for place in free_places:
                    if place != None:
                        free_places = int(place)
                        break
                continue
            if row == 10:
                commercial_places= []
                for col in range(2, sheet.max_column + 1):
                    commercial_places.append(sheet.cell(row, col).value)
                for place in commercial_places:
                    if place != None:
                        commercial_places= int(place)
                        break
                continue

            if sheet.cell(row, 1).value != None and str(sheet.cell(row, 1).value)[0] in '1234567890':
                scan_rate = 1
                if sheet.cell(row, 2).value == snils:
                    found = 1
                    snils_rate = int(sheet.cell(row, 18).value)
                    snils_state = sheet.cell(row, 19).value
                if sheet.cell(row, 18).value != None:
                    rating.append(int(sheet.cell(row, 18).value))
                else:
                    rating.append(0)
    if found:
        rating_with_docs.append(snils_rate)
        rating_with_docs.sort()
        rating_with_docs_rev = list(reversed(rating_with_docs))
        rating.sort(reverse=True)
        position_1 = rating.index(snils_rate)
        rating.reverse()
        position_2 = len(rating) - rating.index(snils_rate)
        pos_docs_1, pos_docs_2 = rating_with_docs.index(snils_rate), len(rating_with_docs)- rating_with_docs_rev.index(snils_rate)
        data = {'program_name': program_name, 'start_position': position_1, 'end_position': position_2, 'free_places': free_places, 'commercial_places': commercial_places,
                'student_points': snils_rate, 'student_program_types': snils_state, 'start_position_docs': pos_docs_1, 'end_position_docs': pos_docs_2}
        wb.close()
        return data
    return